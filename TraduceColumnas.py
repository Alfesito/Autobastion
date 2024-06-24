import openpyxl
from mtranslate import translate
from tqdm import tqdm
import re

# Cargar el archivo de Excel
archivo_excel = r'/Users/andresalfarofernandez/DocumentosPC/VisualStudio_code/Scripts/Autobastion/output.xlsx'
libro = openpyxl.load_workbook(archivo_excel)
hoja = libro.active

# Especificar las columnas que deseas traducir (A, B, C, E, F, G y H)
columnas_a_traducir = ['I']

# Calcular el total de celdas a traducir
total_celdas = (hoja.max_row - 1) * len(columnas_a_traducir)

# Configurar la barra de progreso
barra_progreso = tqdm(total=total_celdas, desc="Progreso")

# Función para encontrar el índice de cierre de la llave
def encontrar_cierre_llave(texto):
    nivel = 0
    for i, char in enumerate(texto):
        if char == '{':
            nivel += 1
        elif char == '}':
            nivel -= 1
            if nivel == 0:
                return i
    return -1

# Función para excluir texto entre comillas simples y dobles y añadir espacios
def excluir_comillas(texto):
    patrones = re.findall(r"'[^']*'|\"[^\"]*\"", texto)
    segmentos = re.split(r"('.*?'|\".*?\")", texto)
    segmentos_con_espacios = []
    for segmento in segmentos:
        if segmento in patrones:
            segmentos_con_espacios.append(' ' + segmento + ' ')
        else:
            segmentos_con_espacios.append(segmento)
    return segmentos_con_espacios, patrones

# Función para recombinar texto con exclusiones
def recombinar_texto(segmentos, patrones, traducciones):
    resultado = []
    i = 0
    for segmento in segmentos:
        if segmento.strip() in patrones:
            resultado.append(segmento)
        else:
            resultado.append(traducciones[i])
            i += 1
    return ''.join(resultado)

# Iterar a través de las filas comenzando desde la fila 2
for fila in range(2, hoja.max_row + 1):
    # Iterar sobre las columnas especificadas
    for columna in columnas_a_traducir:
        # Leer el valor de la celda en la columna actual
        celda = hoja[columna + str(fila)]
        texto_original = celda.value

        # Traducir el texto si no está vacío
        if texto_original and texto_original != "-":
            try:
                # Separar el texto antes y después de "#!/usr/bin/env bash"
                partes = texto_original.split("#!/usr/bin/env bash", 1)
                texto_a_traducir = partes[0]
                
                # Excluir texto entre comillas y añadir espacios
                segmentos, patrones = excluir_comillas(texto_a_traducir)
                
                # Traducir solo los segmentos que no están entre comillas
                traducciones = [translate(segmento, 'es') for segmento in segmentos if segmento.strip() not in patrones]
                
                # Recombinar el texto traducido con las exclusiones
                texto_traducido = recombinar_texto(segmentos, patrones, traducciones)
                
                if len(partes) > 1:
                    # Encontrar el índice de cierre de la llave
                    indice_cierre = encontrar_cierre_llave(partes[1])
                    if (indice_cierre != -1):
                        # Dividir el texto después del bash script
                        script_no_traducir = partes[1][:indice_cierre + 1]
                        resto_traducir = partes[1][indice_cierre + 1:]
                        
                        # Excluir texto entre comillas en el resto
                        segmentos, patrones = excluir_comillas(resto_traducir)
                        
                        # Traducir solo los segmentos que no están entre comillas
                        traducciones = [translate(segmento, 'es') for segmento in segmentos if segmento.strip() not in patrones]
                        
                        # Recombinar el resto del texto traducido
                        resto_traducido = recombinar_texto(segmentos, patrones, traducciones)
                        
                        # Recombinar todo el texto
                        texto_final = texto_traducido + '\n\n' + "#!/usr/bin/env bash" + script_no_traducir + '\n\n' + resto_traducido
                    else:
                        # Si no se encuentra el cierre, no traducir el script
                        texto_final = texto_traducido + '\n\n' + "#!/usr/bin/env bash" + partes[1]
                else:
                    texto_final = texto_traducido

                # Escribir el texto final en la misma celda
                celda.value = texto_final
            except Exception as e:
                print(f"Error al traducir la fila {fila}, columna {columna}: {e}")
        
        # Actualizar la barra de progreso
        barra_progreso.update(1)

# Cerrar la barra de progreso
barra_progreso.close()

# Guardar los cambios en el archivo de Excel
libro.save(r'/Users/andresalfarofernandez/DocumentosPC/VisualStudio_code/Scripts/Autobastion/output_es.xlsx')
