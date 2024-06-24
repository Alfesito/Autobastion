import openpyxl
from tqdm import tqdm
import requests
import json
import time
from openpyxl.styles import PatternFill

# Tu clave de API
api_key = "AIzaSyAkO8l32JaDyHcwxVSE7JutO5-NrYRTLCc"

# URL de la API
url = f"https://generativelanguage.googleapis.com/v1/models/gemini-pro:generateContent?key={api_key}"

# Cargar el archivo de Excel
archivo_excel = r'/Users/andresalfarofernandez/DocumentosPC/VisualStudio_code/Scripts/Autobastion/output.xlsx'
libro = openpyxl.load_workbook(archivo_excel)
hoja = libro.active

# Especificar las columnas que deseas traducir (A, B, C, E, F, G, H, I)
columnas_a_traducir = ['F']

# Calcular el total de celdas a traducir
total_celdas = (hoja.max_row - 1) * len(columnas_a_traducir)

# Configurar la barra de progreso
barra_progreso = tqdm(total=total_celdas, desc="Progreso")

# Definir el color amarillo para resaltar celdas
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Función para realizar la solicitud a la API con reintentos
def traducir_texto(texto_original):
    max_reintentos = 5
    espera_inicial = 1  # 1 segundo
    for intento in range(max_reintentos):
        try:
            # Datos de la solicitud
            data = {
                "contents": [
                    {
                        "role": "user",
                        "parts": [{"text": f"Quiero que la salida sea en texto plano (no quiero que me la des como un .md), traduciendo al español el texto natural y elimina los saltos de linea si procede y que elimines cuando hay un Page y el numero. Pero no los comandos o scripts, estos quiero que me los pongas con sangrado si aplica en el caso: {texto_original}"}]
                    }
                ]
            }

            # Encabezados de la solicitud
            headers = {
                "Content-Type": "application/json"
            }

            # Envío de la solicitud POST a la API
            response = requests.post(url, headers=headers, data=json.dumps(data))
            response.raise_for_status()  # Lanza un error para códigos de estado HTTP malos
            result = response.json()

            # Verificar si hay candidatos y mostrar el contenido
            if "candidates" in result and len(result["candidates"]) > 0:
                content = result["candidates"][0]["content"]["parts"][0]["text"]
                return content
            else:
                print("No valid response received from the API.")
                return None

        except requests.exceptions.HTTPError as e:
            if response.status_code == 429:
                print(f"Intento {intento + 1}/{max_reintentos} - Error 429: Too Many Requests. Esperando {espera_inicial} segundos antes de reintentar.")
                time.sleep(espera_inicial)
                espera_inicial *= 2  # Incrementar el tiempo de espera exponencialmente
            else:
                raise e

    raise Exception(f"Error 429 persistente después de {max_reintentos} reintentos")

# Iterar a través de las filas comenzando desde la fila 2
for fila in range(2, hoja.max_row + 1):
    # Iterar sobre las columnas especificadas
    for columna in columnas_a_traducir:
        # Leer el valor de la celda en la columna actual
        celda = hoja[columna + str(fila)]
        texto_original = celda.value

        # Traducir el texto si no está vacío
        if texto_original and texto_original != "-":
            try:
                content = traducir_texto(texto_original)
                if content:
                    celda.value = content
                else:
                    celda.fill = fill_yellow
            except Exception as e:
                print(f"Error al traducir la fila {fila}, columna {columna}: {e}")
                celda.fill = fill_yellow
        
        # Actualizar la barra de progreso
        barra_progreso.update(1)
        # Pausa breve para evitar problemas de tasa de solicitud
        time.sleep(0.1)  # Ajusta el tiempo de espera según sea necesario

# Cerrar la barra de progreso
barra_progreso.close()

# Guardar los cambios en el archivo de Excel
libro.save(r'/Users/andresalfarofernandez/DocumentosPC/VisualStudio_code/Scripts/Autobastion/output_es.xlsx')
